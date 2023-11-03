from operator import mod
from statistics import mean

import openpyxl
import json

from werkzeug.datastructures import FileStorage

import builtins

def xlsx_to_json(input: FileStorage):
    # Load the Excel workbook
        workbook = openpyxl.load_workbook(input, data_only=True)

        # Select the desired sheet by name
        sheet = workbook[workbook.sheetnames[0]]

        # Create a list to hold the rows as dictionaries
        data = []

        # Iterate through the rows and columns in the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for column, value in enumerate(row):
                header_cell = sheet.cell(row=1, column=column + 1)
                header = header_cell.value
                row_data[header] = str(value)
            data.append(row_data)

        # Convert the list of dictionaries to JSON
        json_data = json.dumps(data, indent=4)

        return json_data


if __name__ == '__main__':
    print('Hello world!')